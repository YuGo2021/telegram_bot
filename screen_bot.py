import telebot

from os import getenv
from selenium import webdriver
from telebot import types
import random
from sys import exit
import time
import openpyxl
from datetime import datetime
import os
from winreg import *
import time
from dotenv import load_dotenv
from selenium.webdriver.chrome.service import Service
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.chrome.options import Options

#s = Service('chromedriver.exe')
s = Service(ChromeDriverManager().install())

time.sleep(10)


# Путь в реестре
key_my = OpenKey(HKEY_CURRENT_USER, r'SOFTWARE\Microsoft\Windows\CurrentVersion\Run', 0, KEY_ALL_ACCESS)
# Установить программу "screen_bot" в автозагрузку
#dir = input("Скопируйте и вставьте путь к файлу бота: ")
#dir_bot = f"{dir}"
#SetValueEx(key_my, 'mynotepad', 0, REG_SZ, dir_bot)
#dir_bot = os.path.abspath('screen_bot.exe')
dir_bot = r"C:\Users\yuri.golubev\Desktop\telegram_bot\screen_bot.exe"
print(dir_bot)
SetValueEx(key_my, 'screen_bot', 0, REG_SZ, dir_bot)
# Закрыть реестр
CloseKey(key_my)



homepath = os.getenv('USERPROFILE')
#print(homepath)  # C:\Users\MyUser

chrome_path = os.path.normpath(homepath + '/AppData/Local/Google/Chrome/User Data/Default')
#print(chrome_path)  # C:\Users\MyUser\Music



#bot_token = getenv("BOT_TOKEN")
load_dotenv("bot_api.env")
bot_token = getenv("TOKEN_BOT")

if not bot_token:
    exit("Error: no token provided")

bot = telebot.TeleBot(bot_token, threaded = False)


#настраиваем браузер для корректной работы в headless режиме

#options = webdriver.ChromeOptions()
options = Options()
#options.add_argument("user-data-dir=C:\\Users\\yuri.golubev\\AppData\\Local\\Google\\Chrome\\User Data\\Default")
#options.add_argument("user-data-dir=C:\\Users\\golub\\AppData\\Local\\Google\\Chrome\\User Data\\Default")
options.add_argument(f"user-data-dir={chrome_path}")

#options.add_argument('--headless')
options.add_argument('--headless=new')
options.add_argument('--disable-gpu')
options.add_argument('--disable-dev-shm-usage')
options.add_argument('--no-sandbox')
options.add_experimental_option("excludeSwitches", ["enable-logging"]) # для исключения ошибки по логированию USD

#options1 = webdriver.ChromeOptions()
options1 = Options()

#options1.add_argument("user-data-dir=C:\\Users\\yuri.golubev\\AppData\\Local\\Google\\Chrome\\User Data\\Default")
options1.add_argument(f"user-data-dir={chrome_path}")


admin = [112533702]


#users.append(admin[0])
# def read_users():
#     wb_users = openpyxl.load_workbook("Пользователи.xlsx")
#     sheet_users = wb_users.active
#     users1 = {}
#     for i in range(1, sheet_users.max_row + 1):
#         users1.update({(sheet_users.cell(row=i, column=1)).value: {"name": (sheet_users.cell(row=i, column=2)).value, "status": (sheet_users.cell(row=i, column=3)).value}})
#
#     users1.update({admin[0]:{"name": "Golubev", "status": 1}})
#     wb_users.save("Пользователи.xlsx")
#     return users1
def users_update(user_dict, user_id, users_name, user_status):
    wb_users = openpyxl.load_workbook("Пользователи.xlsx")
    sheet_users = wb_users.active
    #users2 = {}
    # for i in range(1, sheet_users.max_row + 1):
    #     users2.update({(sheet_users.cell(row=i, column=1)).value: {"name": (sheet_users.cell(row=i, column=2)).value,
    #                                                               "status": (sheet_users.cell(row=i, column=3)).value}})

    #users.update({admin[0]: {"name": "Golubev", "status": 1}})
    #print(users)
    # temp_name = "no_fucking_name"
    # if user_id in users:
    #     temp_name = users[user_id]["name"]
    # if users_name != "noname":
    #     users.update({user_id: {"name": users_name, "status": user_status}})
    # else:
    #     #users.update({user_id: {"status": user_status}})
    #     users.update({user_id: {"name": temp_name, "status": user_status}})
    user_dict.update({user_id: {"name": users_name, "status": user_status}})
    for user in user_dict:
        print(user, user_dict[user]["name"], user_dict[user]["status"])
    line = 0
    for user in user_dict:
        line += 1
        sheet_users.cell(row=line, column=1).value = user
        sheet_users.cell(row=line, column=2).value = user_dict[user]["name"]
        sheet_users.cell(row=line, column=3).value = user_dict[user]["status"]
    wb_users.save("Пользователи.xlsx")

def add_user(user_dict, user_id, users_name, user_status):
    wb_users = openpyxl.load_workbook("Пользователи.xlsx")
    sheet_users = wb_users.active
    user_dict.update({user_id: {"name": users_name, "status": user_status}})
    for user in user_dict:
        print(user, user_dict[user]["name"], user_dict[user]["status"])
    line = 0
    for user in user_dict:
        line += 1
        sheet_users.cell(row=line, column=1).value = user
        sheet_users.cell(row=line, column=2).value = user_dict[user]["name"]
        sheet_users.cell(row=line, column=3).value = user_dict[user]["status"]
    wb_users.save("Пользователи.xlsx")

def set_user_status(user_id, status):
    wb_users = openpyxl.load_workbook("Пользователи.xlsx")
    sheet_users = wb_users.active
    users3 = {}
    for i in range(1, sheet_users.max_row + 1):
        users3.update({(sheet_users.cell(row=i, column=1)).value: {"name": (sheet_users.cell(row=i, column=2)).value,
                                                                  "status": (sheet_users.cell(row=i, column=3)).value}})
    for user in users3:
        if user == user_id:
            users3[user]["status"] = status
    for user in users3:
        print(user, users3[user]["name"], users3[user]["status"])
    wb_users.save("Пользователи.xlsx")


wb_users = openpyxl.load_workbook("Пользователи.xlsx")
sheet_users = wb_users.active
global user1
users1 = {}
for i in range(1, sheet_users.max_row + 1):
    users1.update({(sheet_users.cell(row=i, column=1)).value: {"name": (sheet_users.cell(row=i, column=2)).value, "status": (sheet_users.cell(row=i, column=3)).value}})

users1.update({admin[0]:{"name": "Golubev", "status": 1}})
wb_users.save("Пользователи.xlsx")

#print(users)
# only used for console output now
def listener(messages):
    """
    When new messages arrive TeleBot will call this function.
    """
    for m in messages:
        if m.content_type == 'text':
            # print the sent message to the console
            print(str(m.chat.first_name) + " [" + str(m.chat.id) + "]: " + m.text)

@bot.message_handler(func=lambda message: message.chat.id not in users1)
def some(message):
    # wb_users = openpyxl.load_workbook("Пользователи.xlsx")
    # sheet_users = wb_users.active
    # for i in range(1, sheet_users.max_row + 1):
    #     if sheet_users.cell(row=i, column=1) not in users:
    #         users.append(sheet_users.cell(row=i, column=1))

    users_update(users1, int(message.chat.id), message.chat.first_name, 2)
    # users.update({message.chat.id: {"name": message.chat.first_name, "status": 2}})
    # for user in users:
    #     print(user, users[user]["name"], users[user]["status"])
    # line = 0
    # for user in users:
    #     line += 1
    #     sheet_users.cell(row=line, column=1).value = user
    #     sheet_users.cell(row=line, column=2).value = users[user]["name"]
    #     sheet_users.cell(row=line, column=3).value = users[user]["status"]
    # wb_users.save("Пользователи.xlsx")
    bot.send_message(message.chat.id, 'Это закртый бот. Запросите доступ у администратора')
    bot.send_message(admin[0], 'Есть новые пользователи для подключения. Нажмите /start')
    #
    # markup = types.ReplyKeyboardMarkup(resize_keyboard=True)
    # item0 = types.KeyboardButton("Запросить доступ")
    # markup.add(item0)
    # bot.send_message(message.chat.id, 'Запросить доступ?', reply_markup=markup)
#    bot.send_message(admin[0], 'Запросить доступ', reply_markup=markup)

# Команда start
@bot.message_handler(commands=["start"])
def start(m, res=False):
    # обновляем спсиок users
    # wb_users = openpyxl.load_workbook("Пользователи.xlsx")
    # sheet_users = wb_users.active
    # for i in range(1, sheet_users.max_row + 1):
    #     if sheet_users.cell(row=i, column=1) not in users:
    #         users.append(sheet_users.cell(row=i, column=1))
    #users = read_users()
    users = users1
    if m.chat.id == admin[0]:
        print(users)
        for user in users:
            if users[user]["status"] == 2:
                # markup = types.ReplyKeyboardMarkup(resize_keyboard=True)
                # item_yes = types.KeyboardButton(f"AddUser:{user}")
                # item_no = types.KeyboardButton(f"Decline:{user}")
                # markup.row(item_yes, item_no)
                # bot.send_message(m.chat.id, f'{users[user]["name"]} [{user}] запрашивает доступ к боту', reply_markup=markup)
                keyboard = telebot.types.InlineKeyboardMarkup()
                button1 = telebot.types.InlineKeyboardButton(text="Предоставить доступ", callback_data=f"AddUser:{user}/{users[user]['name']}")
                button2 = telebot.types.InlineKeyboardButton(text="Отказать в доступе", callback_data=f"Decline:{user}/{users[user]['name']}")
                keyboard.row(button1, button2)
                bot.send_message(m.chat.id, f'{users[user]["name"]} [{user}] запрашивает доступ к боту', reply_markup=keyboard)
    if m.chat.id in users and users[m.chat.id]["status"] == 1:
        #bot.send_message(m.chat.id, "Добро пожаловать в бот! Для начала работы нажмите /start")

        # Добавляем кнопки
        markup1 = types.ReplyKeyboardMarkup(resize_keyboard=True)
        item1 = types.KeyboardButton("Общий борд КЦ")
        item2 = types.KeyboardButton("Врачи")
        item3 = types.KeyboardButton("ВО")
        item4 = types.KeyboardButton("ОО")
        item5 = types.KeyboardButton("Дежурный")
        item6 = types.KeyboardButton("B2B")
        # markup.add(item1)
        # markup.add(item2)
        # markup.add(item3)
        # markup.add(item4)
        markup1.row(item1, item5, item2)
        markup1.row(item3, item4, item6)
        bot.send_message(m.chat.id, 'Для получения нужного скриншота нажми на кнопку (время загрузки около 10 сек)',  reply_markup=markup1)
    elif m.chat.id in users and users[m.chat.id]["status"] == 2:
        bot.send_message(m.chat.id, 'Ваш запрос на доступ к боту на рассмотрении. Для ускорения обратитесь к администратору')


# Получение сообщений от юзера
@bot.message_handler(content_types=["text"])
def handle_text(message):
    # wb_users = openpyxl.load_workbook("Пользователи.xlsx")
    # sheet_users = wb_users.active
    # for i in range(1, sheet_users.max_row + 1):
    #     if sheet_users.cell(row=i, column=1) not in users:
    #         users.append(sheet_users.cell(row=i, column=1))
    #users4 = read_users()
    users4 = users1
    print(f"Запрос от {message.chat.first_name} в {datetime.now()}")
    if message.chat.id in users4 and users4[message.chat.id]["status"] == 1:

        # отправляем скрин борда по запросу
        if message.text.strip() == 'Общий борд КЦ' :
            #answer = random.choice(facts)
            uid = message.chat.id
            driver = webdriver.Chrome(options=options, service=s)
            driver.set_window_size(1920, 1080)
            now = datetime.now()
            if 'send_time1' not in globals():
                global send_time1
                send_time1 = datetime(2017, 7, 18, 4, 52, 33, 51204)
            delta = now - send_time1
            if delta.seconds > 60:
                global photo_path1
                photo_path1 = "Image\\" + str(random.randint(10000000, 99999999)) + '.png'

                driver.get("https://grafana2.gemotest.ru:3000/d/i_tq1wY7z/glavnaia-2?orgId=1&refresh=5m")
                #driver.get("https://mail.ru/")
                time.sleep(10)
                driver.save_screenshot(photo_path1)
                #global send_time1
                send_time1 = datetime.now()

            bot.send_photo(uid, photo=open(photo_path1, 'rb'))

            driver.quit()
            #os.remove(photo_path)

        elif message.text.strip() == 'Врачи':
            uid = message.chat.id
            driver = webdriver.Chrome(options=options, service=s)
            driver.set_window_size(1920, 1080)
            now = datetime.now()
            if 'send_time2' not in globals():
                global send_time2
                send_time2 = datetime(2017, 7, 18, 4, 52, 33, 51204)

            delta = now - send_time2
            if delta.seconds > 60:
                global photo_path2
                photo_path2 = "Image\\" + str(random.randint(10000000, 99999999)) + '.png'

                driver.get("https://grafana2.gemotest.ru:3000/d/WCQ6wTsnz/vrachi?orgId=1&refresh=1m")
                #driver.get("https://yandex.ru/")
                time.sleep(10)
                driver.save_screenshot(photo_path2)
                send_time2 = datetime.now()
            bot.send_photo(uid, photo=open(photo_path2, 'rb'))
            driver.quit()
            #os.remove(photo_path)
                #answer = random.choice(thinks)
        elif message.text.strip() == 'ВО':
            uid = message.chat.id
            driver = webdriver.Chrome(options=options, service=s)
            driver.set_window_size(1920, 1080)
            now = datetime.now()
            if 'send_time3' not in globals():
                global send_time3
                send_time3 = datetime(2017, 7, 18, 4, 52, 33, 51204)
            delta = now - send_time3
            if delta.seconds > 60:
                global photo_path3
                photo_path3 = "Image\\" + str(random.randint(10000000, 99999999)) + '.png'

                driver.get("https://grafana2.gemotest.ru:3000/d/n_7NiLrnk/bvo?orgId=1&refresh=15m")
                #driver.get("https://yandex.ru/")
                time.sleep(10)
                driver.save_screenshot(photo_path3)
                send_time3 = datetime.now()
            bot.send_photo(uid, photo=open(photo_path3, 'rb'))
            driver.quit()
        elif message.text.strip() == 'ОО':
            uid = message.chat.id
            driver = webdriver.Chrome(options=options, service=s)
            driver.set_window_size(1920, 1080)
            now = datetime.now()
            if 'send_time4' not in globals():
                global send_time4
                send_time4 = datetime(2017, 7, 18, 4, 52, 33, 51204)
            delta = now - send_time4
            if delta.seconds > 60:
                global photo_path4
                photo_path4 = "Image\\" + str(random.randint(10000000, 99999999)) + '.png'

                driver.get("https://grafana2.gemotest.ru:3000/d/-zNr5Zq7z/oo?orgId=1&refresh=1m")
                #driver.get("https://yandex.ru/")
                time.sleep(10)
                driver.save_screenshot(photo_path4)
                send_time4 = datetime.now()
            bot.send_photo(uid, photo=open(photo_path4, 'rb'))
            driver.quit()
        elif message.text.strip() == 'Дежурный':
            uid = message.chat.id
            driver = webdriver.Chrome(options=options, service=s)
            driver.set_window_size(1920, 1200)
            now = datetime.now()
            if 'send_time5' not in globals():
                global send_time5
                send_time5 = datetime(2017, 7, 18, 4, 52, 33, 51204)
            delta = now - send_time5
            if delta.seconds > 60:
                global photo_path5
                photo_path5 = "Image\\" + str(random.randint(10000000, 99999999)) + '.png'

                driver.get("https://grafana2.gemotest.ru:3000/d/_xv1gmGVk/dezhurnyi?orgId=1&refresh=30s")
                # driver.get("https://yandex.ru/")
                time.sleep(10)
                driver.save_screenshot(photo_path5)
                send_time5 = datetime.now()
            bot.send_photo(uid, photo=open(photo_path5, 'rb'))
            driver.quit()
        elif message.text.strip() == 'B2B':
            uid = message.chat.id
            driver = webdriver.Chrome(options=options, service=s)
            driver.set_window_size(1920, 1200)
            now = datetime.now()
            if 'send_time6' not in globals():
                global send_time6
                send_time6 = datetime(2017, 7, 18, 4, 52, 33, 51204)
            delta = now - send_time6
            if delta.seconds > 60:
                global photo_path6
                photo_path6 = "Image\\" + str(random.randint(10000000, 99999999)) + '.png'

                driver.get("https://grafana2.gemotest.ru:3000/d/Dlb_TeI4k/b2b?orgId=1")
                # driver.get("https://yandex.ru/")
                time.sleep(10)
                driver.save_screenshot(photo_path6)
                send_time6 = datetime.now()
            bot.send_photo(uid, photo=open(photo_path6, 'rb'))
            driver.quit()
    #    if message.chat.id == admin[0]:
    #       if message.text.strip() == 'отладка':
        elif message.text.strip() == 'отладка':
            uid = message.chat.id
            photo_path = "Image\\" + str(random.randint(10000000, 99999999)) + '.png'
            driver = webdriver.Chrome(options=options1, service=s)
            driver.set_window_size(1920, 1080)
            driver.get("https://grafana2.gemotest.ru:3000/d/i_tq1wY7z/glavnaia-2?orgId=1&refresh=5m")
            #driver.get("https://yandex.ru/")
            time.sleep(30)
            driver.save_screenshot(photo_path)
            bot.send_photo(uid, photo=open(photo_path, 'rb'))
            driver.quit()
        elif message.text.strip() == 'тест':
            uid = message.chat.id
            photo_path = "Image\\" + str(random.randint(10000000, 99999999)) + '.png'
            driver = webdriver.Chrome(options=options, service=s)
            driver.set_window_size(1920, 1080)
            #driver.get("https://grafana2.gemotest.ru:3000/d/i_tq1wY7z/glavnaia-2?orgId=1&refresh=5m")
            driver.get("https://yandex.ru/")
            time.sleep(10)
            driver.save_screenshot(photo_path)
            bot.send_photo(uid, photo=open(photo_path, 'rb'))
            driver.quit()
        elif message.text.strip()[:4] == 'All:':
            for user in users4:
                if users4[user]["status"] == 1:
                    bot.send_message(user, f"{message.text.strip()[5:]}")
        # if message.text.strip()[:7] == 'AddUser:':
        #     print(message.text.strip[8:])
        #     if message.chat.id == admin[0]:
        #         for user in users:
        #             if user == message.text.strip[8:]:
        #                 bot.send_message(message.chat.id, 'Пользовтель уже добавлен')
        #                 break
        #         else:
        #             users[message.text.strip[8:]]["status"] = 1
        #             for user in users:
        #                 print(user, users[user]["name"], users[user]["status"])
        #             line = 0
        #             for i in users:
        #                 line += 1
        #                 sheet_users.cell(row=line, column=1).value = i
        #                 sheet_users.cell(row=line, column=2).value = users[i]["name"]
        #                 sheet_users.cell(row=line, column=3).value = users[i]["status"]
        #             wb_users.save("Пользователи.xlsx")
        #             bot.send_message(message.chat.id, 'Пользовтель добавлен')
        #             bot.send_message(message.text.strip[8:], 'Вам предоставлен доступ к боту')
    # if message.chat.id not in users:
    #     if message.text.strip() == 'Запросить доступ':
    #
    #         uid = message.chat.id
    #         users.update({uid:{"name":message.chat.first_name, "status":2}})
    #         for user in users:
    #             print(user, user["name"], user["status"])
    #         line = 0
    #         for user in users:
    #             line += 1
    #             sheet_users.cell(row=line, column=1).value = user
    #             sheet_users.cell(row=line, column=2).value = user["name"]
    #             sheet_users.cell(row=line, column=3).value = user["status"]
    #         wb_users.save("Пользователи.xlsx")
    #
    #
    #
    #         bot.send_message(admin[0], f'{message.chat.first_name} [{uid}] запрашивает доступ к боту')
    #         markup2 = types.ReplyKeyboardMarkup(resize_keyboard=True)
    #         item5 = types.KeyboardButton(f'AddUser:{uid}')
    #         item6 = types.KeyboardButton(f"Decline:{uid}")
    #
    #         markup2.row(item5, item6)
    #         bot.send_message(admin[0],
    #                          f'{message.chat.first_name} [{uid}] запрашивает доступ к боту',  reply_markup=markup2)

@bot.callback_query_handler(func=lambda call: True)
def callback_function1(callback_obj):

    #users5 = read_users()
    command = callback_obj.data.partition("/")
    if command[0][:8] == "AddUser:":
        for user in users1:
            if user == int(command[0][8:]) and users1[user]["status"] == 1:
                #bot.send_message(message.chat.id, 'Пользовтель уже добавлен')
                bot.send_message(callback_obj.from_user.id, f"Пользовтель {callback_obj.data[8:]} уже добавлен")
                break

        else:
            #set_user_status(int(command[0][8:]), 1)
            #users_update(int(command[0][8:]), command[2], 1)
            add_user(users1, int(command[0][8:]), command[2], 1)
            #users_update(int(callback_obj.data[8:]), "noname", 1)

            #users_update(callback_obj.data[8:], "noname", 1)
            # users[int(callback_obj.data[8:])]["status"] = 1
            # for user in users:
            #     print(user, users[user]["name"], users[user]["status"])
            # line = 0
            # for i in users:
            #     line += 1
            #     sheet_users.cell(row=line, column=1).value = i
            #     sheet_users.cell(row=line, column=2).value = users[i]["name"]
            #     sheet_users.cell(row=line, column=3).value = users[i]["status"]
            # wb_users.save("Пользователи.xlsx")
            #bot.send_message(message.chat.id, 'Пользовтель добавлен')
            bot.send_message(callback_obj.from_user.id, f"Пользовтель {callback_obj.data[8:]} добавлен")
            bot.send_message(int(command[0][8:]), 'Вам предоставлен доступ к боту. Нажмите /start')
            #break
    elif callback_obj.data[:8] == "Decline:":
        #set_user_status(int(command[0][8:]), 0)
        #users_update(int(command[0][8:]), command[2], 0)
        add_user(users1, int(command[0][8:]), command[2], 0)
        #users_update(int(callback_obj.data[8:]), "noname", 0)
        # users[int(callback_obj.data[8:])]["status"] = 0
        # # print(users, users["status"])
        # line = 0
        # for i in users:
        #     line += 1
        #     sheet_users.cell(row=line, column=1).value = i
        #     sheet_users.cell(row=line, column=2).value = users[i]["name"]
        #     sheet_users.cell(row=line, column=3).value = users[i]["status"]
        # wb_users.save("Пользователи.xlsx")
        bot.send_message(int(command[0][8:]), 'Вам отказано в достпуе к боту')

        bot.send_message(callback_obj.from_user.id, f"Пользователь {callback_obj.data[8:]} заблокирован")


    bot.answer_callback_query(callback_query_id=callback_obj.id)

#Handles all text messages that match the regular expression
# @bot.message_handler(regexp=r"((AddUser:)(\d*))")
# def handle_message(message):
#     print(message.text.strip[8:])
#     if message.chat.id == admin[0]:
#         for user in users:
#             if user == message.text.strip[8:]:
#                 bot.send_message(message.chat.id,'Пользовтель уже добавлен')
#                 break
#         else:
#             users[message.text.strip[8:]]["status"] = 1
#             for user in users:
#                 print(user, users[user]["name"], users[user]["status"])
#             line = 0
#             for i in users:
#                 line += 1
#                 sheet_users.cell(row=line, column=1).value = i
#                 sheet_users.cell(row=line, column=2).value = users[i]["name"]
#                 sheet_users.cell(row=line, column=3).value = users[i]["status"]
#             wb_users.save("Пользователи.xlsx")
#             bot.send_message(message.chat.id, 'Пользовтель добавлен')
#             bot.send_message(message.text.strip[8:], 'Вам предоставлен доступ к боту')

# @bot.message_handler(regexp=r"((Decline:)(\d*)")
# def handle_message(message):
#     if message.chat.id in admin:
#         users[message.text.strip[8:]]["status"] = 0
#         #print(users, users["status"])
#         line = 0
#         for i in users:
#             line += 1
#             sheet_users.cell(row=line, column=1).value = i
#             sheet_users.cell(row=line, column=2).value = users[i]["name"]
#             sheet_users.cell(row=line, column=3).value = users[i]["status"]
#         wb_users.save("Пользователи.xlsx")
#         bot.send_message(message.text.strip[8:], 'Вам отказано в достпуе к боту')

# Запускаем бота
while True:
    try:
        bot.polling(none_stop=True, interval=0)
        break
    except Exception as e:
        print(f"Ошибка {e} ({datetime.now()})")
        bot.stop_polling()
        time.sleep(20)



# отправка картинки пользователю
# uid = message.chat.id
# photo_path = str(uid) + '.png'
# driver = webdriver.Chrome(chrome_options = options)
# driver.set_window_size(1280, 720)
# driver.get(url)
# driver.save_screenshot(photo_path)
# bot.send_photo(uid, photo = open(photo_path, 'rb'))
# driver.quit()
# os.remove(photo_path)